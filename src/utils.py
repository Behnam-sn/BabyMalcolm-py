import json
from os import mkdir
from openpyxl import load_workbook

import keys
from item import Item


def excel_to_items(excel_file_path: str) -> list[Item]:
    items = []
    item_last_id = 0

    workbook = load_workbook(
        data_only=True,
        filename=excel_file_path,
    )
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  # type: ignore
        item_last_id += 1

        item = Item(
            id=item_last_id,
            date=row[keys.ITEM_DATE],
            day_in_week=row[keys.ITEM_DAY_IN_WEEK],
            start_time=row[keys.ITEM_START_TIME],
            category=row[keys.ITEM_CATEGORY],
            subject=row[keys.ITEM_SUBJECT],
            detail=row[keys.ITEM_DETAIL],
            time_spent=row[keys.ITEM_TIME_SPENT],
        )

        items.append(item)

    return items


def item_to_dictionary(items: list[Item]) -> dict:
    dictionary = {}

    for item in items:
        dictionary[item.id] = item.convert_to_dictionary()

    return dictionary


def create_folder(path: str, name: str) -> None:
    try:
        mkdir(f"{path}\\{name}")

    except Exception:
        pass


def dictionary_to_json_file(items: dict, file_name: str) -> None:
    create_folder(path=".", name="dist")

    with open(f"dist\\{file_name}.json", "w", encoding="utf-8") as file:
        json.dump(items, file, ensure_ascii=False, indent=4)


def dictionary_to_text_file(items: dict, file_name: str) -> None:
    create_folder(path=".\\", name="dist")

    text_file = open(f"dist\\{file_name}.txt", "w", encoding="utf-8")

    for category in items:
        for subject in items[category]:
            hour = items[category][subject]["hour"]
            minute = items[category][subject]["minute"]

            stringified_hour = number_to_string(hour)
            stringified_minute = number_to_string(minute)

            time_spent = f"{stringified_hour}:{stringified_minute}"

            text_file.write(f"{category}\t{subject}\t{time_spent}\n")


def filter_items_by_date(items: list[Item], start_int: int, end_int: int) -> list[Item]:
    start = number_to_string(start_int)
    end = number_to_string(end_int)

    filtered_items = filter(
        lambda item: (start <= item.get_day() and end >= item.get_day()),
        items,
    )

    return list(filtered_items)


def number_to_string(number: int) -> str:
    string = str(number)

    if len(string) == 1:
        string = f"0{string}"

    return string


def new_dictionary_with_defualt_categories() -> dict:
    dictionary = {}

    for category in keys.CATEGORIES:
        dictionary[category] = {}

    return dictionary


def summarize_items(items: list[Item]) -> dict:
    summary = {}

    for item in items:
        if item.category not in summary:
            summary[item.category] = {}

        if item.subject is not None and item.subject not in summary[item.category]:
            summary[item.category][item.subject] = {"hour": 0, "minute": 0}

        if (
            item.subject is not None
            and summary[item.category][item.subject] is not None
            and item.time_spent is not None
        ):
            summary[item.category][item.subject]["hour"] += item.time_spent.hour
            summary[item.category][item.subject]["minute"] += item.time_spent.minute

            if summary[item.category][item.subject]["minute"] >= 60:
                hour = int(summary[item.category][item.subject]["minute"] / 60)
                summary[item.category][item.subject]["hour"] += hour
                summary[item.category][item.subject]["minute"] = (
                    summary[item.category][item.subject]["minute"] % 60
                )

    return summary


def sort_summary(summary: dict) -> dict:
    sorted_summary = new_dictionary_with_defualt_categories()

    for category in summary:
        if category not in sorted_summary:
            sorted_summary[category] = {}

        subjects = summary[category]
        # sorted_subjects = dict(sorted(subjects.items()))
        # sorted_summary[category] = sorted_subjects
        sorted_summary[category] = subjects

    return sorted_summary
